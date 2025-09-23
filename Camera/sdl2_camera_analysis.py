#!/usr/bin/env python3
# SDL2 Camera Analysis - Lightweight version for embedded systems with Weston

import os
import sys
import ctypes
import glob
import subprocess
import time
import re
from datetime import datetime

# Check if required modules are available
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Required module missing: {e}")
    print("Please install required modules:")
    print("pip install pandas openpyxl")
    sys.exit(1)

try:
    import pygame
    import gi
    gi.require_version("Gst", "1.0")
    from gi.repository import Gst
except ImportError as e:
    print(f"Required module missing: {e}")
    print("Please install required modules:")
    print("pip install pygame")
    print("GStreamer python bindings required")
    sys.exit(1)

class SDL2CameraAnalyzer:
    def __init__(self):
        # Set environment variables for Docker/Weston compatibility
        os.environ['SDL_VIDEODRIVER'] = 'wayland'
        os.environ['XDG_RUNTIME_DIR'] = '/tmp/1000-runtime-dir'
        os.environ['WAYLAND_DISPLAY'] = 'wayland-0'

        # Disable audio to avoid ALSA/PulseAudio issues in Docker
        os.environ['SDL_AUDIODRIVER'] = 'dummy'

        # Initialize pygame without audio
        pygame.mixer.pre_init(frequency=0, size=0, channels=0, buffer=0)
        pygame.mixer.quit()
        pygame.init()
        pygame.mixer.quit()  # Ensure audio is disabled

        Gst.init(None)

        # Display settings for rotated 1280x800 screen
        self.screen_width = 800
        self.screen_height = 1280

        # Create display with fallback options
        try:
            # Try fullscreen first
            self.screen = pygame.display.set_mode((self.screen_width, self.screen_height), pygame.FULLSCREEN)
            print("Created fullscreen display")
        except Exception as e:
            print(f"Fullscreen failed: {e}, trying windowed mode")
            try:
                # Try windowed mode
                self.screen = pygame.display.set_mode((self.screen_width, self.screen_height))
                print("Created windowed display")
            except Exception as e:
                print(f"Windowed mode failed: {e}, trying smaller window")
                # Try smaller window as last resort
                self.screen_width = 640
                self.screen_height = 480
                self.screen = pygame.display.set_mode((self.screen_width, self.screen_height))
                print("Created fallback small display")

        pygame.display.set_caption("SDL2 Camera Analysis")

        # Colors
        self.BLACK = (0, 0, 0)
        self.WHITE = (255, 255, 255)
        self.BLUE = (68, 114, 196)
        self.GREEN = (198, 239, 206)
        self.RED = (255, 199, 206)
        self.GRAY = (128, 128, 128)
        self.LIGHT_GRAY = (240, 240, 240)

        # Fonts
        self.font_large = pygame.font.Font(None, 36)
        self.font_medium = pygame.font.Font(None, 24)
        self.font_small = pygame.font.Font(None, 18)

        # Analysis settings
        self.recording_duration = 15  # seconds
        self.wait_duration = 16  # seconds
        self.temp_dir = "temp_analysis"
        self.output_excel = "real_camera_analysis_sdl2.xlsx"

        # Analysis state
        self.video_devices = []
        self.analysis_results = {}
        self.current_test = {}
        self.total_combinations = 0
        self.completed_combinations = 0
        self.is_analyzing = False
        self.is_recording = False
        self.pipeline = None

        # UI state
        self.scroll_offset = 0
        self.max_scroll = 0

        # UI elements (buttons)
        self.start_button = pygame.Rect(50, 100, 300, 80)
        self.exit_button = pygame.Rect(450, 100, 200, 80)
        self.scroll_up_button = pygame.Rect(700, 200, 80, 40)
        self.scroll_down_button = pygame.Rect(700, 1200, 80, 40)

        # Find video devices and their capabilities
        print("Scanning video devices for capabilities...")
        self.get_real_device_capabilities()
        self.create_temp_directory()

        # Start main loop
        self.running = True
        self.clock = pygame.time.Clock()

    def parse_v4l2_output(self, device_path):
        """Parse v4l2-ctl output to extract real device capabilities"""
        try:
            result = subprocess.run(['v4l2-ctl', '--device', device_path, '--list-formats-ext'],
                                  capture_output=True, text=True, timeout=5)

            if result.returncode != 0 or not result.stdout.strip():
                return {}

            capabilities = {}
            current_format = None

            lines = result.stdout.split('\n')

            for line in lines:
                line = line.strip()

                # Look for format lines
                format_match = re.search(r"\[(\d+)\]:\s+'([^']+)'\s+\(([^)]+)\)", line)
                if format_match:
                    format_code = format_match.group(2)
                    format_desc = format_match.group(3)
                    current_format = format_code
                    capabilities[current_format] = {
                        'description': format_desc,
                        'resolutions': {}
                    }
                    continue

                # Look for size lines
                size_match = re.search(r"Size:\s+Discrete\s+(\d+)x(\d+)", line)
                if size_match and current_format:
                    width = int(size_match.group(1))
                    height = int(size_match.group(2))
                    resolution = (width, height)

                    if resolution not in capabilities[current_format]['resolutions']:
                        capabilities[current_format]['resolutions'][resolution] = []
                    continue

                # Look for interval lines
                interval_match = re.search(r"Interval:\s+Discrete\s+[\d.]+s\s+\(([\d.]+)\s+fps\)", line)
                if interval_match and current_format:
                    fps = float(interval_match.group(1))
                    # Add this fps to the last resolution found
                    resolutions = capabilities[current_format]['resolutions']
                    if resolutions:
                        last_resolution = list(resolutions.keys())[-1]
                        capabilities[current_format]['resolutions'][last_resolution].append(fps)

            return capabilities

        except Exception as e:
            print(f"Error parsing v4l2 output for {device_path}: {e}")
            return {}

    def get_real_device_capabilities(self):
        """Get video devices and their REAL capabilities from v4l2-ctl"""
        for device_path in glob.glob('/dev/video*'):
            try:
                print(f"Checking {device_path}...")
                capabilities = self.parse_v4l2_output(device_path)

                if capabilities:
                    device_info = {
                        'path': device_path,
                        'capabilities': capabilities
                    }
                    self.video_devices.append(device_info)

                    # Count total combinations
                    for format_name, format_data in capabilities.items():
                        for resolution, fps_list in format_data['resolutions'].items():
                            self.total_combinations += len(fps_list)

            except Exception as e:
                print(f"Error checking {device_path}: {e}")
                continue

        print(f"Found {len(self.video_devices)} usable video devices")
        print(f"Total combinations to test: {self.total_combinations}")

    def create_temp_directory(self):
        """Create temporary directory for test recordings"""
        try:
            if not os.path.exists(self.temp_dir):
                os.makedirs(self.temp_dir)
                print(f"Created temporary directory: {self.temp_dir}")
        except Exception as e:
            print(f"Error creating temp directory: {e}")

    def draw_button(self, rect, text, color, text_color=None, enabled=True):
        """Draw a button with text"""
        if not enabled:
            color = self.GRAY

        pygame.draw.rect(self.screen, color, rect)
        pygame.draw.rect(self.screen, self.BLACK, rect, 2)

        if text_color is None:
            text_color = self.WHITE if color == self.BLUE else self.BLACK

        text_surface = self.font_medium.render(text, True, text_color)
        text_rect = text_surface.get_rect(center=rect.center)
        self.screen.blit(text_surface, text_rect)

    def draw_text(self, text, x, y, font, color=None):
        """Draw text at position"""
        if color is None:
            color = self.BLACK
        text_surface = font.render(text, True, color)
        self.screen.blit(text_surface, (x, y))
        return text_surface.get_height()

    def draw_wrapped_text(self, text, x, y, max_width, font, color=None):
        """Draw text with word wrapping"""
        if color is None:
            color = self.BLACK

        words = text.split(' ')
        lines = []
        current_line = ""

        for word in words:
            test_line = current_line + word + " "
            if font.size(test_line)[0] <= max_width:
                current_line = test_line
            else:
                if current_line:
                    lines.append(current_line.strip())
                current_line = word + " "

        if current_line:
            lines.append(current_line.strip())

        total_height = 0
        for line in lines:
            height = self.draw_text(line, x, y + total_height, font, color)
            total_height += height + 2

        return total_height

    def draw_device_info(self, y_offset):
        """Draw device information"""
        if not self.video_devices:
            return self.draw_text("No video devices found", 50, y_offset, self.font_medium, self.RED)

        current_y = y_offset

        for device_info in self.video_devices:
            device_path = device_info['path']
            capabilities = device_info['capabilities']

            # Device title
            current_y += self.draw_text(f"Device: {device_path}", 50, current_y, self.font_medium, self.BLUE) + 10

            for fmt, fmt_data in capabilities.items():
                combinations = sum(len(fps_list) for fps_list in fmt_data['resolutions'].values())
                current_y += self.draw_text(f"  • {fmt}: {len(fmt_data['resolutions'])} resolutions, {combinations} combinations",
                                          70, current_y, self.font_small) + 5

                for resolution, fps_list in fmt_data['resolutions'].items():
                    w, h = resolution
                    fps_str = ', '.join([f"{fps:.0f}" for fps in sorted(fps_list)])
                    current_y += self.draw_text(f"    {w}x{h}: {fps_str} fps",
                                              90, current_y, self.font_small, self.GRAY) + 3

            current_y += 15

        return current_y - y_offset

    def draw_progress(self, y_offset):
        """Draw analysis progress"""
        if not self.is_analyzing:
            return 0

        # Progress bar
        progress_pct = (self.completed_combinations / self.total_combinations) * 100 if self.total_combinations > 0 else 0
        progress_width = int((self.screen_width - 100) * (progress_pct / 100))

        # Background
        pygame.draw.rect(self.screen, self.LIGHT_GRAY, (50, y_offset, self.screen_width - 100, 30))
        # Progress
        pygame.draw.rect(self.screen, self.GREEN, (50, y_offset, progress_width, 30))
        # Border
        pygame.draw.rect(self.screen, self.BLACK, (50, y_offset, self.screen_width - 100, 30), 2)

        # Progress text
        progress_text = f"{self.completed_combinations}/{self.total_combinations} ({progress_pct:.1f}%)"
        text_surface = self.font_small.render(progress_text, True, self.BLACK)
        text_rect = text_surface.get_rect(center=(self.screen_width // 2, y_offset + 15))
        self.screen.blit(text_surface, text_rect)

        current_y = y_offset + 40

        # Current test info
        if hasattr(self, 'current_test') and self.current_test:
            device_path = self.current_test.get('device_path', '')
            format_name = self.current_test.get('format', '')
            resolution = self.current_test.get('resolution', (0, 0))
            fps = self.current_test.get('fps', 0)

            w, h = resolution
            test_info = f"Testing: {device_path} {format_name} {w}x{h}@{fps}fps"
            current_y += self.draw_text(test_info, 50, current_y, self.font_small, self.BLUE) + 10

        return current_y - y_offset

    def handle_scroll(self, event):
        """Handle scroll events"""
        if event.type == pygame.MOUSEBUTTONDOWN:
            if self.scroll_up_button.collidepoint(event.pos):
                self.scroll_offset = max(0, self.scroll_offset - 50)
            elif self.scroll_down_button.collidepoint(event.pos):
                self.scroll_offset = min(self.max_scroll, self.scroll_offset + 50)

    def handle_events(self):
        """Handle SDL2 events"""
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                self.running = False

            elif event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE:
                    self.running = False
                elif event.key == pygame.K_UP:
                    self.scroll_offset = max(0, self.scroll_offset - 50)
                elif event.key == pygame.K_DOWN:
                    self.scroll_offset = min(self.max_scroll, self.scroll_offset + 50)

            elif event.type == pygame.MOUSEBUTTONDOWN:
                if self.start_button.collidepoint(event.pos) and not self.is_analyzing:
                    self.start_complete_analysis()
                elif self.exit_button.collidepoint(event.pos):
                    self.running = False
                else:
                    self.handle_scroll(event)

    def draw_ui(self):
        """Draw the main UI"""
        self.screen.fill(self.WHITE)

        # Title
        title = "SDL2 Camera Analysis"
        title_surface = self.font_large.render(title, True, self.BLUE)
        title_rect = title_surface.get_rect(center=(self.screen_width // 2, 40))
        self.screen.blit(title_surface, title_rect)

        # Subtitle
        subtitle = f"Real-time analysis of {self.total_combinations} combinations"
        subtitle_surface = self.font_medium.render(subtitle, True, self.BLACK)
        subtitle_rect = subtitle_surface.get_rect(center=(self.screen_width // 2, 70))
        self.screen.blit(subtitle_surface, subtitle_rect)

        # Buttons
        start_text = "⏳ Running..." if self.is_analyzing else "🚀 Start Analysis"
        self.draw_button(self.start_button, start_text, self.BLUE if not self.is_analyzing else self.GRAY,
                        enabled=not self.is_analyzing)
        self.draw_button(self.exit_button, "Exit", self.RED)

        # Progress section
        current_y = 200
        if self.is_analyzing:
            current_y += self.draw_progress(current_y) + 20

        # Scrollable content area
        content_y = current_y - self.scroll_offset

        # Device info
        device_info_height = self.draw_device_info(content_y)

        # Update max scroll
        total_content_height = device_info_height + 200
        visible_height = self.screen_height - current_y - 100
        self.max_scroll = max(0, total_content_height - visible_height)

        # Scroll buttons
        if self.max_scroll > 0:
            self.draw_button(self.scroll_up_button, "↑", self.LIGHT_GRAY, enabled=self.scroll_offset > 0)
            self.draw_button(self.scroll_down_button, "↓", self.LIGHT_GRAY, enabled=self.scroll_offset < self.max_scroll)

        pygame.display.flip()

    def start_complete_analysis(self):
        """Start the complete automated analysis"""
        if self.is_analyzing:
            return

        self.is_analyzing = True
        self.completed_combinations = 0
        self.analysis_results = {}

        print("Starting complete analysis...")

        # Start analysis in next frame
        pygame.time.set_timer(pygame.USEREVENT + 1, 100)  # Start in 100ms

    def run_next_test(self):
        """Run the next test in sequence"""
        if not self.is_analyzing:
            return

        # Find next combination to test
        next_test = self.get_next_test_combination()

        if not next_test:
            # All tests complete
            self.complete_analysis()
            return

        self.current_test = next_test
        device_path = next_test['device_path']
        format_name = next_test['format']
        resolution = next_test['resolution']
        fps = next_test['fps']

        w, h = resolution
        print(f"Testing: {device_path} {format_name} {w}x{h}@{fps}fps")

        # Start recording this combination
        self.record_test_video()

    def get_next_test_combination(self):
        """Get the next combination to test"""
        current_index = 0

        for device_info in self.video_devices:
            device_path = device_info['path']
            capabilities = device_info['capabilities']

            for format_name, format_data in capabilities.items():
                for resolution, fps_list in format_data['resolutions'].items():
                    for fps in sorted(fps_list):
                        if current_index == self.completed_combinations:
                            return {
                                'device_path': device_path,
                                'format': format_name,
                                'resolution': resolution,
                                'fps': fps
                            }
                        current_index += 1

        return None  # No more combinations

    def record_test_video(self):
        """Record a test video for the current combination"""
        try:
            device_path = self.current_test['device_path']
            format_name = self.current_test['format']
            resolution = self.current_test['resolution']
            fps = self.current_test['fps']

            w, h = resolution

            # Generate temp filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if format_name == 'H264':
                ext = 'mp4'
            else:
                ext = 'avi'

            filename = f"test_{format_name}_{w}x{h}_{fps:.0f}fps_{timestamp}.{ext}"
            output_file = os.path.join(self.temp_dir, filename)

            # Build recording pipeline
            if format_name == 'H264':
                caps = f"video/x-h264,width={w},height={h},framerate={fps:.0f}/1"
                pipeline_str = f"v4l2src device={device_path} ! {caps} ! h264parse ! mp4mux ! filesink location={output_file}"

            elif format_name == 'MJPG':
                caps = f"image/jpeg,width={w},height={h},framerate={fps:.0f}/1"
                pipeline_str = f"v4l2src device={device_path} ! {caps} ! avimux ! filesink location={output_file}"

            else:  # YUYV
                caps = f"video/x-raw,format=YUY2,width={w},height={h},framerate={fps:.0f}/1"
                pipeline_str = f"v4l2src device={device_path} ! {caps} ! videoconvert ! x264enc ! avimux ! filesink location={output_file}"

            # Try to create and start pipeline
            self.pipeline = Gst.parse_launch(pipeline_str)
            ret = self.pipeline.set_state(Gst.State.PLAYING)

            self.is_recording = True
            self.current_test['output_file'] = output_file
            self.current_test['record_start'] = time.time()

            # Schedule finish after wait_duration
            pygame.time.set_timer(pygame.USEREVENT + 2, self.wait_duration * 1000)

        except Exception as e:
            print(f"Recording error: {e}")
            # Simple cleanup and mark as failed
            if self.pipeline:
                try:
                    self.pipeline.set_state(Gst.State.NULL)
                    self.pipeline = None
                except:
                    pass

            self.record_test_result(False, 0, 0)
            self.completed_combinations += 1
            # Schedule next test
            pygame.time.set_timer(pygame.USEREVENT + 1, 100)

    def finish_test_recording(self):
        """Finish recording and measure results - simple approach"""
        try:
            if self.pipeline:
                # Send EOS to properly finish the file
                self.pipeline.send_event(Gst.Event.new_eos())
                time.sleep(0.5)  # Wait for EOS to process
                self.pipeline.set_state(Gst.State.NULL)
                self.pipeline = None

            self.is_recording = False

            output_file = self.current_test['output_file']

            # Check if file was created and measure it
            if os.path.exists(output_file):
                file_size = os.path.getsize(output_file)
                file_size_mb = file_size / (1024 * 1024)

                # Simple check: if file exists with any reasonable size, consider success
                if file_size > 1024:  # At least 1KB
                    # Calculate actual bitrate based on 15-second recording
                    bits = file_size * 8
                    bitrate_bps = bits / self.recording_duration
                    bitrate_kbps = bitrate_bps / 1000

                    print(f"Success: {output_file} - {file_size_mb:.2f} MB, {bitrate_kbps:.1f} kbps")
                    self.record_test_result(True, file_size_mb, bitrate_kbps)
                else:
                    print(f"Failed: {output_file} too small ({file_size} bytes)")
                    self.record_test_result(False, 0, 0)

                # Always delete the file to save space
                try:
                    os.remove(output_file)
                except:
                    pass

            else:
                print(f"Failed: {output_file} not created")
                self.record_test_result(False, 0, 0)

        except Exception as e:
            print(f"Error finishing recording: {e}")
            self.record_test_result(False, 0, 0)

        self.completed_combinations += 1

        # Schedule next test
        pygame.time.set_timer(pygame.USEREVENT + 1, 500)

    def record_test_result(self, success, file_size_mb, bitrate_kbps):
        """Record the result of a test"""
        device_path = self.current_test['device_path']
        format_name = self.current_test['format']
        resolution = self.current_test['resolution']
        fps = self.current_test['fps']

        # Store result
        if device_path not in self.analysis_results:
            self.analysis_results[device_path] = {}

        if format_name not in self.analysis_results[device_path]:
            self.analysis_results[device_path][format_name] = []

        self.analysis_results[device_path][format_name].append({
            'resolution': resolution,
            'fps': fps,
            'success': success,
            'file_size_mb': file_size_mb,
            'bitrate_kbps': bitrate_kbps
        })

    def complete_analysis(self):
        """Complete the analysis and generate Excel file"""
        self.is_analyzing = False
        print("Analysis complete! Generating Excel file...")

        # Generate Excel file
        try:
            self.generate_real_excel_file()
            print(f"Excel file saved: {self.output_excel}")
        except Exception as e:
            print(f"Excel generation failed: {e}")

        # Cleanup temp directory
        try:
            os.rmdir(self.temp_dir)
        except:
            pass

    def generate_real_excel_file(self):
        """Generate Excel file with real measured data"""
        print(f"Generating Excel file: {self.output_excel}")

        wb = Workbook()
        wb.remove(wb.active)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Process each device/format combination
        for device_path, device_data in self.analysis_results.items():
            device_name = device_path.replace('/dev/', '')

            for format_name, format_results in device_data.items():
                if not format_results:
                    continue

                # Create worksheet
                sheet_name = f"{device_name}_{format_name}"
                ws = wb.create_sheet(title=sheet_name)

                # Convert to DataFrame
                df_data = []
                for result in format_results:
                    w, h = result['resolution']
                    df_data.append({
                        'Resolution': f"{w}x{h}",
                        'Width': w,
                        'Height': h,
                        'FPS': result['fps'],
                        'Real Bitrate (kbps)': round(result['bitrate_kbps'], 1) if result['success'] else 0,
                        'Real File Size 15s (MB)': round(result['file_size_mb'], 3) if result['success'] else 0,
                        'Works': "✓" if result['success'] else "✗"
                    })

                df = pd.DataFrame(df_data)

                # Write title
                ws['A1'] = f"SDL2 REAL DATA: {device_path} - {format_name}"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:H1')

                # Create matrix
                resolutions = df['Resolution'].unique()
                fps_values = sorted(df['FPS'].unique())

                # Headers
                row = 3
                ws[f'A{row}'] = "Resolution"
                ws[f'A{row}'].font = header_font
                ws[f'A{row}'].fill = header_fill
                ws[f'A{row}'].alignment = center_align
                ws[f'A{row}'].border = border

                col = 2
                for fps in fps_values:
                    cell = ws.cell(row=row, column=col, value=f"{fps} FPS")
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border
                    col += 1

                # Fill matrix
                row += 1
                for resolution in resolutions:
                    ws.cell(row=row, column=1, value=resolution).font = Font(bold=True)
                    ws.cell(row=row, column=1).alignment = center_align
                    ws.cell(row=row, column=1).border = border

                    col = 2
                    for fps in fps_values:
                        matching = df[(df['Resolution'] == resolution) & (df['FPS'] == fps)]

                        if not matching.empty:
                            data = matching.iloc[0]
                            bitrate = data['Real Bitrate (kbps)']
                            filesize = data['Real File Size 15s (MB)']
                            works = data['Works']

                            if works == "✓":
                                cell_content = f"{bitrate} kbps\n{filesize} MB\n✓ SDL2"
                                cell_fill = success_fill
                            else:
                                cell_content = "FAILED\n0 MB\n✗"
                                cell_fill = fail_fill

                            cell = ws.cell(row=row, column=col, value=cell_content)
                            cell.alignment = center_align
                            cell.border = border
                            cell.fill = cell_fill

                        else:
                            cell = ws.cell(row=row, column=col, value="N/A")
                            cell.alignment = center_align
                            cell.border = border

                        col += 1
                    row += 1

                # Add detailed table
                row += 2
                ws.cell(row=row, column=1, value="SDL2 REAL MEASURED DATA:").font = Font(bold=True, size=12)
                row += 1

                headers = ['Resolution', 'FPS', 'Real Bitrate (kbps)', 'Real File Size 15s (MB)', 'Works']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border

                row += 1
                for _, data in df.iterrows():
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=data[header])
                        cell.alignment = center_align
                        cell.border = border

                        if header == 'Works':
                            if data[header] == "✓":
                                cell.fill = success_fill
                            else:
                                cell.fill = fail_fill
                    row += 1

                # Auto-adjust column widths
                for col_num in range(1, ws.max_column + 1):
                    max_length = 0
                    column_letter = get_column_letter(col_num)

                    for row_num in range(1, ws.max_row + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 20)
                    ws.column_dimensions[column_letter].width = adjusted_width

        # Create summary sheet
        summary_ws = wb.create_sheet(title="SDL2_REAL_Summary", index=0)
        summary_ws['A1'] = "SDL2 Real Camera Analysis Summary"
        summary_ws['A1'].font = Font(bold=True, size=16)

        row = 3
        total_tested = 0
        total_successful = 0

        for device_path, device_data in self.analysis_results.items():
            summary_ws.cell(row=row, column=1, value=f"Device: {device_path}").font = Font(bold=True)
            row += 1

            for format_name, format_results in device_data.items():
                successful = len([r for r in format_results if r['success']])
                total = len(format_results)

                total_tested += total
                total_successful += successful

                summary_ws.cell(row=row, column=2, value=f"{format_name}: {successful}/{total} combinations successful")
                row += 1
            row += 1

        summary_ws.cell(row=row, column=1, value=f"TOTAL: {total_successful}/{total_tested} combinations successful").font = Font(bold=True, size=14)

        wb.save(self.output_excel)
        print(f"SDL2 analysis Excel file saved: {self.output_excel}")

    def run(self):
        """Main application loop"""
        while self.running:
            # Handle timer events
            for event in pygame.event.get():
                if event.type == pygame.USEREVENT + 1:
                    # Run next test
                    self.run_next_test()
                elif event.type == pygame.USEREVENT + 2:
                    # Finish current recording
                    self.finish_test_recording()
                else:
                    # Put the event back for normal handling
                    pygame.event.post(event)
                    break

            self.handle_events()
            self.draw_ui()
            self.clock.tick(30)  # 30 FPS

        pygame.quit()

if __name__ == "__main__":
    try:
        app = SDL2CameraAnalyzer()
        app.run()
    except KeyboardInterrupt:
        print("\nApplication interrupted by user")
    except Exception as e:
        print(f"Application error: {e}")
    finally:
        pygame.quit()