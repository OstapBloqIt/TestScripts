#!/usr/bin/env python3
# SDL2 Camera Analysis - Safe version for Docker/Weston environments

import os
import sys
import ctypes
import glob
import subprocess
import time
import re
from datetime import datetime

# Check display before importing heavy modules
def check_display():
    """Check if display is available"""
    try:
        # Check for Wayland display
        if os.environ.get('WAYLAND_DISPLAY'):
            print(f"Wayland display found: {os.environ.get('WAYLAND_DISPLAY')}")
            return True
        # Check for X11 display
        elif os.environ.get('DISPLAY'):
            print(f"X11 display found: {os.environ.get('DISPLAY')}")
            return True
        else:
            print("No display environment found")
            return False
    except Exception as e:
        print(f"Display check failed: {e}")
        return False

if not check_display():
    print("Error: No display available. Make sure you're running in a graphical environment.")
    sys.exit(1)

# Check if required modules are available
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Required module missing: {e}")
    print("Please install required modules:")
    print("pip install pandas openpyxl")
    sys.exit(1)

# Set environment variables before importing pygame/gi
os.environ['XDG_RUNTIME_DIR'] = '/tmp/1000-runtime-dir'
os.environ['WAYLAND_DISPLAY'] = 'wayland-0'
os.environ['SDL_AUDIODRIVER'] = 'dummy'

try:
    import pygame
    import gi
    gi.require_version("Gst", "1.0")
    from gi.repository import Gst
except ImportError as e:
    print(f"Required module missing: {e}")
    print("Please install required modules:")
    print("pip install pygame")
    print("GStreamer python bindings required")
    sys.exit(1)

class SafeSDL2CameraAnalyzer:
    def __init__(self):
        print("Initializing SDL2 Camera Analyzer...")

        # Disable pygame audio completely
        os.environ['SDL_AUDIODRIVER'] = 'dummy'

        # Initialize pygame with minimal setup
        try:
            pygame.mixer.pre_init(frequency=0, size=0, channels=0, buffer=0)
            pygame.mixer.quit()

            # Only init video and events
            pygame.display.init()
            pygame.font.init()

            print("Pygame initialized successfully")
        except Exception as e:
            print(f"Pygame initialization failed: {e}")
            sys.exit(1)

        # Initialize GStreamer
        try:
            Gst.init(None)
            print("GStreamer initialized successfully")
        except Exception as e:
            print(f"GStreamer initialization failed: {e}")

        # Display settings
        self.screen_width = 800
        self.screen_height = 600  # Start with smaller size

        # Create display
        self.create_display()

        # Colors
        self.BLACK = (0, 0, 0)
        self.WHITE = (255, 255, 255)
        self.BLUE = (68, 114, 196)
        self.GREEN = (198, 239, 206)
        self.RED = (255, 199, 206)
        self.GRAY = (128, 128, 128)
        self.LIGHT_GRAY = (240, 240, 240)

        # Fonts - use None for default system font
        try:
            self.font_large = pygame.font.Font(None, 36)
            self.font_medium = pygame.font.Font(None, 24)
            self.font_small = pygame.font.Font(None, 18)
            print("Fonts initialized successfully")
        except Exception as e:
            print(f"Font initialization failed: {e}")
            # Use basic font as fallback
            self.font_large = pygame.font.SysFont('arial', 36)
            self.font_medium = pygame.font.SysFont('arial', 24)
            self.font_small = pygame.font.SysFont('arial', 18)

        # Analysis settings
        self.recording_duration = 15  # seconds
        self.wait_duration = 16  # seconds
        self.temp_dir = "temp_analysis"
        self.output_excel = "real_camera_analysis_sdl2.xlsx"

        # Analysis state
        self.video_devices = []
        self.analysis_results = {}
        self.current_test = {}
        self.total_combinations = 0
        self.completed_combinations = 0
        self.is_analyzing = False
        self.is_recording = False
        self.pipeline = None

        # UI state
        self.scroll_offset = 0
        self.max_scroll = 0

        # UI elements (buttons) - adjusted for smaller screen
        button_width = min(250, self.screen_width // 3)
        button_height = 60

        self.start_button = pygame.Rect(50, 100, button_width, button_height)
        self.exit_button = pygame.Rect(self.screen_width - 150, 100, 100, button_height)
        self.scroll_up_button = pygame.Rect(self.screen_width - 80, 200, 60, 30)
        self.scroll_down_button = pygame.Rect(self.screen_width - 80, self.screen_height - 80, 60, 30)

        # Initialize application
        self.initialize_app()

        # Start main loop
        self.running = True
        self.clock = pygame.time.Clock()

    def create_display(self):
        """Create display with progressive fallback"""
        print("Creating display...")

        display_modes = [
            (800, 1280, pygame.FULLSCREEN, "fullscreen 800x1280"),
            (800, 600, pygame.FULLSCREEN, "fullscreen 800x600"),
            (800, 1280, 0, "windowed 800x1280"),
            (800, 600, 0, "windowed 800x600"),
            (640, 480, 0, "windowed 640x480"),
        ]

        for width, height, flags, description in display_modes:
            try:
                print(f"Trying {description}...")
                self.screen = pygame.display.set_mode((width, height), flags)
                self.screen_width = width
                self.screen_height = height
                pygame.display.set_caption("SDL2 Camera Analysis")
                print(f"Successfully created {description}")
                return
            except Exception as e:
                print(f"Failed to create {description}: {e}")
                continue

        raise Exception("Could not create any display mode")

    def initialize_app(self):
        """Initialize application components"""
        print("Finding video devices and their capabilities...")
        self.get_real_device_capabilities()
        self.create_temp_directory()

    def parse_v4l2_output(self, device_path):
        """Parse v4l2-ctl output to extract real device capabilities"""
        try:
            result = subprocess.run(['v4l2-ctl', '--device', device_path, '--list-formats-ext'],
                                  capture_output=True, text=True, timeout=5)

            if result.returncode != 0 or not result.stdout.strip():
                return {}

            capabilities = {}
            current_format = None

            lines = result.stdout.split('\n')

            for line in lines:
                line = line.strip()

                # Look for format lines
                format_match = re.search(r"\[(\d+)\]:\s+'([^']+)'\s+\(([^)]+)\)", line)
                if format_match:
                    format_code = format_match.group(2)
                    format_desc = format_match.group(3)
                    current_format = format_code
                    capabilities[current_format] = {
                        'description': format_desc,
                        'resolutions': {}
                    }
                    continue

                # Look for size lines
                size_match = re.search(r"Size:\s+Discrete\s+(\d+)x(\d+)", line)
                if size_match and current_format:
                    width = int(size_match.group(1))
                    height = int(size_match.group(2))
                    resolution = (width, height)

                    if resolution not in capabilities[current_format]['resolutions']:
                        capabilities[current_format]['resolutions'][resolution] = []
                    continue

                # Look for interval lines
                interval_match = re.search(r"Interval:\s+Discrete\s+[\d.]+s\s+\(([\d.]+)\s+fps\)", line)
                if interval_match and current_format:
                    fps = float(interval_match.group(1))
                    # Add this fps to the last resolution found
                    resolutions = capabilities[current_format]['resolutions']
                    if resolutions:
                        last_resolution = list(resolutions.keys())[-1]
                        capabilities[current_format]['resolutions'][last_resolution].append(fps)

            return capabilities

        except Exception as e:
            print(f"Error parsing v4l2 output for {device_path}: {e}")
            return {}

    def get_real_device_capabilities(self):
        """Get video devices and their REAL capabilities from v4l2-ctl"""
        for device_path in glob.glob('/dev/video*'):
            try:
                print(f"Checking {device_path}...")
                capabilities = self.parse_v4l2_output(device_path)

                if capabilities:
                    device_info = {
                        'path': device_path,
                        'capabilities': capabilities
                    }
                    self.video_devices.append(device_info)

                    # Count total combinations
                    for format_name, format_data in capabilities.items():
                        for resolution, fps_list in format_data['resolutions'].items():
                            self.total_combinations += len(fps_list)

            except Exception as e:
                print(f"Error checking {device_path}: {e}")
                continue

        print(f"Found {len(self.video_devices)} usable video devices")
        print(f"Total combinations to test: {self.total_combinations}")

    def create_temp_directory(self):
        """Create temporary directory for test recordings"""
        try:
            if not os.path.exists(self.temp_dir):
                os.makedirs(self.temp_dir)
                print(f"Created temporary directory: {self.temp_dir}")
        except Exception as e:
            print(f"Error creating temp directory: {e}")

    def draw_button(self, rect, text, color, text_color=None, enabled=True):
        """Draw a button with text"""
        try:
            if not enabled:
                color = self.GRAY

            pygame.draw.rect(self.screen, color, rect)
            pygame.draw.rect(self.screen, self.BLACK, rect, 2)

            if text_color is None:
                text_color = self.WHITE if color == self.BLUE else self.BLACK

            text_surface = self.font_medium.render(text, True, text_color)
            text_rect = text_surface.get_rect(center=rect.center)
            self.screen.blit(text_surface, text_rect)
        except Exception as e:
            print(f"Error drawing button: {e}")

    def draw_text(self, text, x, y, font, color=None):
        """Draw text at position"""
        try:
            if color is None:
                color = self.BLACK
            text_surface = font.render(str(text), True, color)
            self.screen.blit(text_surface, (x, y))
            return text_surface.get_height()
        except Exception as e:
            print(f"Error drawing text: {e}")
            return 20

    def draw_ui(self):
        """Draw the main UI"""
        try:
            self.screen.fill(self.WHITE)

            # Title
            title = "SDL2 Camera Analysis"
            title_surface = self.font_large.render(title, True, self.BLUE)
            title_rect = title_surface.get_rect(center=(self.screen_width // 2, 40))
            self.screen.blit(title_surface, title_rect)

            # Subtitle
            subtitle = f"Real-time analysis of {self.total_combinations} combinations"
            subtitle_surface = self.font_medium.render(subtitle, True, self.BLACK)
            subtitle_rect = subtitle_surface.get_rect(center=(self.screen_width // 2, 70))
            self.screen.blit(subtitle_surface, subtitle_rect)

            # Buttons
            start_text = "Running..." if self.is_analyzing else "Start Analysis"
            self.draw_button(self.start_button, start_text, self.BLUE if not self.is_analyzing else self.GRAY,
                            enabled=not self.is_analyzing)
            self.draw_button(self.exit_button, "Exit", self.RED)

            # Progress section
            if self.is_analyzing:
                self.draw_progress(200)

            # Device info
            self.draw_device_info(300)

            pygame.display.flip()

        except Exception as e:
            print(f"Error drawing UI: {e}")

    def draw_progress(self, y_offset):
        """Draw analysis progress"""
        try:
            if not self.is_analyzing:
                return

            # Progress bar
            progress_pct = (self.completed_combinations / self.total_combinations) * 100 if self.total_combinations > 0 else 0
            progress_width = int((self.screen_width - 100) * (progress_pct / 100))

            # Background
            pygame.draw.rect(self.screen, self.LIGHT_GRAY, (50, y_offset, self.screen_width - 100, 20))
            # Progress
            pygame.draw.rect(self.screen, self.GREEN, (50, y_offset, progress_width, 20))
            # Border
            pygame.draw.rect(self.screen, self.BLACK, (50, y_offset, self.screen_width - 100, 20), 2)

            # Progress text
            progress_text = f"{self.completed_combinations}/{self.total_combinations} ({progress_pct:.1f}%)"
            self.draw_text(progress_text, 50, y_offset + 25, self.font_small)

            # Current test info
            if hasattr(self, 'current_test') and self.current_test:
                device_path = self.current_test.get('device_path', '')
                format_name = self.current_test.get('format', '')
                resolution = self.current_test.get('resolution', (0, 0))
                fps = self.current_test.get('fps', 0)

                w, h = resolution
                test_info = f"Testing: {device_path} {format_name} {w}x{h}@{fps}fps"
                self.draw_text(test_info, 50, y_offset + 45, self.font_small, self.BLUE)

        except Exception as e:
            print(f"Error drawing progress: {e}")

    def draw_device_info(self, y_offset):
        """Draw device information"""
        try:
            if not self.video_devices:
                self.draw_text("No video devices found", 50, y_offset, self.font_medium, self.RED)
                return

            current_y = y_offset

            for device_info in self.video_devices:
                device_path = device_info['path']
                capabilities = device_info['capabilities']

                # Device title
                current_y += self.draw_text(f"Device: {device_path}", 50, current_y, self.font_medium, self.BLUE) + 10

                for fmt, fmt_data in capabilities.items():
                    combinations = sum(len(fps_list) for fps_list in fmt_data['resolutions'].values())
                    current_y += self.draw_text(f"  {fmt}: {len(fmt_data['resolutions'])} resolutions, {combinations} combinations",
                                              70, current_y, self.font_small) + 5

                current_y += 15

        except Exception as e:
            print(f"Error drawing device info: {e}")

    def handle_events(self):
        """Handle SDL2 events"""
        try:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    self.running = False

                elif event.type == pygame.KEYDOWN:
                    if event.key == pygame.K_ESCAPE:
                        self.running = False

                elif event.type == pygame.MOUSEBUTTONDOWN:
                    if self.start_button.collidepoint(event.pos) and not self.is_analyzing:
                        self.start_complete_analysis()
                    elif self.exit_button.collidepoint(event.pos):
                        self.running = False

                elif event.type == pygame.USEREVENT + 1:
                    # Run next test
                    self.run_next_test()
                elif event.type == pygame.USEREVENT + 2:
                    # Finish current recording
                    self.finish_test_recording()

        except Exception as e:
            print(f"Error handling events: {e}")

    def start_complete_analysis(self):
        """Start the complete automated analysis"""
        if self.is_analyzing:
            return

        self.is_analyzing = True
        self.completed_combinations = 0
        self.analysis_results = {}

        print("Starting complete analysis...")

        # Start analysis in next frame
        pygame.time.set_timer(pygame.USEREVENT + 1, 100)  # Start in 100ms

    def run_next_test(self):
        """Run the next test in sequence"""
        if not self.is_analyzing:
            return

        # Find next combination to test
        next_test = self.get_next_test_combination()

        if not next_test:
            # All tests complete
            self.complete_analysis()
            return

        self.current_test = next_test
        device_path = next_test['device_path']
        format_name = next_test['format']
        resolution = next_test['resolution']
        fps = next_test['fps']

        w, h = resolution
        print(f"Testing: {device_path} {format_name} {w}x{h}@{fps}fps")

        # Start recording this combination
        self.record_test_video()

    def get_next_test_combination(self):
        """Get the next combination to test"""
        current_index = 0

        for device_info in self.video_devices:
            device_path = device_info['path']
            capabilities = device_info['capabilities']

            for format_name, format_data in capabilities.items():
                for resolution, fps_list in format_data['resolutions'].items():
                    for fps in sorted(fps_list):
                        if current_index == self.completed_combinations:
                            return {
                                'device_path': device_path,
                                'format': format_name,
                                'resolution': resolution,
                                'fps': fps
                            }
                        current_index += 1

        return None  # No more combinations

    def record_test_video(self):
        """Record a test video for the current combination"""
        try:
            device_path = self.current_test['device_path']
            format_name = self.current_test['format']
            resolution = self.current_test['resolution']
            fps = self.current_test['fps']

            w, h = resolution

            # Generate temp filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if format_name == 'H264':
                ext = 'mp4'
            else:
                ext = 'avi'

            filename = f"test_{format_name}_{w}x{h}_{fps:.0f}fps_{timestamp}.{ext}"
            output_file = os.path.join(self.temp_dir, filename)

            # Build recording pipeline
            if format_name == 'H264':
                caps = f"video/x-h264,width={w},height={h},framerate={fps:.0f}/1"
                pipeline_str = f"v4l2src device={device_path} ! {caps} ! h264parse ! mp4mux ! filesink location={output_file}"

            elif format_name == 'MJPG':
                caps = f"image/jpeg,width={w},height={h},framerate={fps:.0f}/1"
                pipeline_str = f"v4l2src device={device_path} ! {caps} ! avimux ! filesink location={output_file}"

            else:  # YUYV
                caps = f"video/x-raw,format=YUY2,width={w},height={h},framerate={fps:.0f}/1"
                pipeline_str = f"v4l2src device={device_path} ! {caps} ! videoconvert ! x264enc ! avimux ! filesink location={output_file}"

            # Try to create and start pipeline
            self.pipeline = Gst.parse_launch(pipeline_str)
            ret = self.pipeline.set_state(Gst.State.PLAYING)

            self.is_recording = True
            self.current_test['output_file'] = output_file
            self.current_test['record_start'] = time.time()

            # Schedule finish after wait_duration
            pygame.time.set_timer(pygame.USEREVENT + 2, self.wait_duration * 1000)

        except Exception as e:
            print(f"Recording error: {e}")
            # Simple cleanup and mark as failed
            if self.pipeline:
                try:
                    self.pipeline.set_state(Gst.State.NULL)
                    self.pipeline = None
                except:
                    pass

            self.record_test_result(False, 0, 0)
            self.completed_combinations += 1
            # Schedule next test
            pygame.time.set_timer(pygame.USEREVENT + 1, 100)

    def finish_test_recording(self):
        """Finish recording and measure results - simple approach"""
        try:
            if self.pipeline:
                # Send EOS to properly finish the file
                self.pipeline.send_event(Gst.Event.new_eos())
                time.sleep(0.5)  # Wait for EOS to process
                self.pipeline.set_state(Gst.State.NULL)
                self.pipeline = None

            self.is_recording = False

            output_file = self.current_test['output_file']

            # Check if file was created and measure it
            if os.path.exists(output_file):
                file_size = os.path.getsize(output_file)
                file_size_mb = file_size / (1024 * 1024)

                # Simple check: if file exists with any reasonable size, consider success
                if file_size > 1024:  # At least 1KB
                    # Calculate actual bitrate based on 15-second recording
                    bits = file_size * 8
                    bitrate_bps = bits / self.recording_duration
                    bitrate_kbps = bitrate_bps / 1000

                    print(f"Success: {output_file} - {file_size_mb:.2f} MB, {bitrate_kbps:.1f} kbps")
                    self.record_test_result(True, file_size_mb, bitrate_kbps)
                else:
                    print(f"Failed: {output_file} too small ({file_size} bytes)")
                    self.record_test_result(False, 0, 0)

                # Always delete the file to save space
                try:
                    os.remove(output_file)
                except:
                    pass

            else:
                print(f"Failed: {output_file} not created")
                self.record_test_result(False, 0, 0)

        except Exception as e:
            print(f"Error finishing recording: {e}")
            self.record_test_result(False, 0, 0)

        self.completed_combinations += 1

        # Schedule next test
        pygame.time.set_timer(pygame.USEREVENT + 1, 500)

    def record_test_result(self, success, file_size_mb, bitrate_kbps):
        """Record the result of a test"""
        device_path = self.current_test['device_path']
        format_name = self.current_test['format']
        resolution = self.current_test['resolution']
        fps = self.current_test['fps']

        # Store result
        if device_path not in self.analysis_results:
            self.analysis_results[device_path] = {}

        if format_name not in self.analysis_results[device_path]:
            self.analysis_results[device_path][format_name] = []

        self.analysis_results[device_path][format_name].append({
            'resolution': resolution,
            'fps': fps,
            'success': success,
            'file_size_mb': file_size_mb,
            'bitrate_kbps': bitrate_kbps
        })

    def complete_analysis(self):
        """Complete the analysis and generate Excel file"""
        self.is_analyzing = False
        print("Analysis complete! Generating Excel file...")

        # Generate Excel file
        try:
            self.generate_real_excel_file()
            print(f"Excel file saved: {self.output_excel}")
        except Exception as e:
            print(f"Excel generation failed: {e}")

        # Cleanup temp directory
        try:
            os.rmdir(self.temp_dir)
        except:
            pass

    def generate_real_excel_file(self):
        """Generate Excel file with real measured data"""
        print(f"Generating Excel file: {self.output_excel}")

        wb = Workbook()
        wb.remove(wb.active)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Process each device/format combination
        for device_path, device_data in self.analysis_results.items():
            device_name = device_path.replace('/dev/', '')

            for format_name, format_results in device_data.items():
                if not format_results:
                    continue

                # Create worksheet
                sheet_name = f"{device_name}_{format_name}"
                ws = wb.create_sheet(title=sheet_name)

                # Convert to DataFrame
                df_data = []
                for result in format_results:
                    w, h = result['resolution']
                    df_data.append({
                        'Resolution': f"{w}x{h}",
                        'Width': w,
                        'Height': h,
                        'FPS': result['fps'],
                        'Real Bitrate (kbps)': round(result['bitrate_kbps'], 1) if result['success'] else 0,
                        'Real File Size 15s (MB)': round(result['file_size_mb'], 3) if result['success'] else 0,
                        'Works': "✓" if result['success'] else "✗"
                    })

                df = pd.DataFrame(df_data)

                # Write title
                ws['A1'] = f"SAFE SDL2 REAL DATA: {device_path} - {format_name}"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:H1')

                # Simple data table (skip complex matrix for now)
                row = 3
                headers = ['Resolution', 'FPS', 'Real Bitrate (kbps)', 'Real File Size 15s (MB)', 'Works']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border

                row += 1
                for _, data in df.iterrows():
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=data[header])
                        cell.alignment = center_align
                        cell.border = border

                        if header == 'Works':
                            if data[header] == "✓":
                                cell.fill = success_fill
                            else:
                                cell.fill = fail_fill
                    row += 1

        # Create summary sheet
        summary_ws = wb.create_sheet(title="SAFE_SDL2_Summary", index=0)
        summary_ws['A1'] = "Safe SDL2 Real Camera Analysis Summary"
        summary_ws['A1'].font = Font(bold=True, size=16)

        row = 3
        total_tested = 0
        total_successful = 0

        for device_path, device_data in self.analysis_results.items():
            summary_ws.cell(row=row, column=1, value=f"Device: {device_path}").font = Font(bold=True)
            row += 1

            for format_name, format_results in device_data.items():
                successful = len([r for r in format_results if r['success']])
                total = len(format_results)

                total_tested += total
                total_successful += successful

                summary_ws.cell(row=row, column=2, value=f"{format_name}: {successful}/{total} combinations successful")
                row += 1
            row += 1

        summary_ws.cell(row=row, column=1, value=f"TOTAL: {total_successful}/{total_tested} combinations successful").font = Font(bold=True, size=14)

        wb.save(self.output_excel)
        print(f"Safe SDL2 analysis Excel file saved: {self.output_excel}")

    def run(self):
        """Main application loop"""
        print("Starting main application loop...")
        try:
            while self.running:
                self.handle_events()
                self.draw_ui()
                self.clock.tick(30)  # 30 FPS

        except Exception as e:
            print(f"Error in main loop: {e}")
        finally:
            print("Shutting down...")
            pygame.quit()

if __name__ == "__main__":
    try:
        print("Starting Safe SDL2 Camera Analyzer...")
        app = SafeSDL2CameraAnalyzer()
        app.run()
    except KeyboardInterrupt:
        print("\nApplication interrupted by user")
    except Exception as e:
        print(f"Application error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        try:
            pygame.quit()
        except:
            pass