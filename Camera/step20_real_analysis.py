#!/usr/bin/env python3
# Step 20: Automated Real Camera Analysis - Records and measures all combinations

import gi
import glob
import subprocess
import time
import re
import os
import sys
from datetime import datetime

# Check if required modules are available
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Required module missing: {e}")
    print("Please install required modules:")
    print("pip install pandas openpyxl")
    sys.exit(1)

gi.require_version("Gtk", "3.0")
gi.require_version("Gst", "1.0")
from gi.repository import Gtk, Gst, GLib

class RealCameraAnalyzer(Gtk.Window):
    def __init__(self):
        super().__init__()
        self.set_title("Step 20: Real Camera Analysis")

        # Fixed window size for rotated display (like working versions)
        self.set_default_size(800, 1280)
        self.set_position(Gtk.WindowPosition.CENTER)
        self.set_resizable(False)

        self.connect("destroy", Gtk.main_quit)

        # Initialize GStreamer
        Gst.init(None)
        self.pipeline = None
        self.is_recording = False
        self.is_analyzing = False

        # Analysis settings - simple approach
        self.recording_duration = 15  # seconds
        self.wait_duration = 16  # seconds - wait for recording to complete
        self.temp_dir = "temp_analysis"
        self.output_excel = "real_camera_analysis.xlsx"

        # Analysis data
        self.video_devices = []
        self.analysis_results = {}
        self.current_test = {}
        self.total_combinations = 0
        self.completed_combinations = 0

        # UI elements for progress
        self.progress_info = None
        self.start_analysis_btn = None

        # Find video devices and their capabilities
        print("Scanning video devices for capabilities...")
        self.get_real_device_capabilities()

        self.setup_layout()
        self.create_temp_directory()

    def parse_v4l2_output(self, device_path):
        """Parse v4l2-ctl output to extract real device capabilities"""
        try:
            result = subprocess.run(['v4l2-ctl', '--device', device_path, '--list-formats-ext'],
                                  capture_output=True, text=True, timeout=5)

            if result.returncode != 0 or not result.stdout.strip():
                return {}

            capabilities = {}
            current_format = None

            lines = result.stdout.split('\n')

            for line in lines:
                line = line.strip()

                # Look for format lines
                format_match = re.search(r"\[(\d+)\]:\s+'([^']+)'\s+\(([^)]+)\)", line)
                if format_match:
                    format_code = format_match.group(2)
                    format_desc = format_match.group(3)
                    current_format = format_code
                    capabilities[current_format] = {
                        'description': format_desc,
                        'resolutions': {}
                    }
                    continue

                # Look for size lines
                size_match = re.search(r"Size:\s+Discrete\s+(\d+)x(\d+)", line)
                if size_match and current_format:
                    width = int(size_match.group(1))
                    height = int(size_match.group(2))
                    resolution = (width, height)

                    if resolution not in capabilities[current_format]['resolutions']:
                        capabilities[current_format]['resolutions'][resolution] = []
                    continue

                # Look for interval lines
                interval_match = re.search(r"Interval:\s+Discrete\s+[\d.]+s\s+\(([\d.]+)\s+fps\)", line)
                if interval_match and current_format:
                    fps = float(interval_match.group(1))
                    # Add this fps to the last resolution found
                    resolutions = capabilities[current_format]['resolutions']
                    if resolutions:
                        last_resolution = list(resolutions.keys())[-1]
                        capabilities[current_format]['resolutions'][last_resolution].append(fps)

            return capabilities

        except Exception as e:
            print(f"Error parsing v4l2 output for {device_path}: {e}")
            return {}

    def get_real_device_capabilities(self):
        """Get video devices and their REAL capabilities from v4l2-ctl"""
        for device_path in glob.glob('/dev/video*'):
            try:
                print(f"Checking {device_path}...")
                capabilities = self.parse_v4l2_output(device_path)

                if capabilities:
                    device_info = {
                        'path': device_path,
                        'capabilities': capabilities
                    }
                    self.video_devices.append(device_info)

                    # Count total combinations
                    for format_name, format_data in capabilities.items():
                        for resolution, fps_list in format_data['resolutions'].items():
                            self.total_combinations += len(fps_list)

            except Exception as e:
                print(f"Error checking {device_path}: {e}")
                continue

        print(f"Found {len(self.video_devices)} usable video devices")
        print(f"Total combinations to test: {self.total_combinations}")

    def create_temp_directory(self):
        """Create temporary directory for test recordings"""
        try:
            if not os.path.exists(self.temp_dir):
                os.makedirs(self.temp_dir)
                print(f"Created temporary directory: {self.temp_dir}")
        except Exception as e:
            print(f"Error creating temp directory: {e}")

    def setup_layout(self):
        # Main vertical box
        main_vbox = Gtk.VBox(spacing=0)
        self.add(main_vbox)

        # Title area
        title_frame = Gtk.Frame()
        title_frame.set_size_request(800, 100)
        title_frame.set_shadow_type(Gtk.ShadowType.IN)
        main_vbox.pack_start(title_frame, False, False, 0)

        title_vbox = Gtk.VBox(spacing=10)
        title_vbox.set_property("margin", 20)
        title_frame.add(title_vbox)

        title_label = Gtk.Label()
        title_label.set_markup("<span font='24'><b>Real Camera Analysis</b></span>")
        title_vbox.pack_start(title_label, False, False, 0)

        subtitle_label = Gtk.Label()
        subtitle_label.set_markup(f"<span font='16'>Records and measures all {self.total_combinations} format/resolution/fps combinations</span>")
        title_vbox.pack_start(subtitle_label, False, False, 0)

        # Analysis controls
        controls_frame = Gtk.Frame()
        controls_frame.set_size_request(800, 150)
        controls_frame.set_shadow_type(Gtk.ShadowType.IN)
        main_vbox.pack_start(controls_frame, False, False, 0)

        controls_vbox = Gtk.VBox(spacing=15)
        controls_vbox.set_property("margin", 20)
        controls_frame.add(controls_vbox)

        # Start analysis button
        button_hbox = Gtk.HBox(spacing=20)
        controls_vbox.pack_start(button_hbox, False, False, 0)

        self.start_analysis_btn = Gtk.Button(label="🚀 Start Complete Analysis")
        self.start_analysis_btn.set_size_request(300, 60)
        self.start_analysis_btn.connect("clicked", self.start_complete_analysis)
        button_hbox.pack_start(self.start_analysis_btn, False, False, 0)

        exit_btn = Gtk.Button(label="Exit")
        exit_btn.set_size_request(100, 60)
        exit_btn.connect("clicked", lambda *_: Gtk.main_quit())
        button_hbox.pack_start(exit_btn, False, False, 0)

        # Warning label
        warning_label = Gtk.Label()
        warning_label.set_markup(f"<span font='12' color='red'><b>⚠️  This will take approximately {self.total_combinations * self.wait_duration} seconds ({(self.total_combinations * self.wait_duration) // 60} minutes)</b></span>")
        controls_vbox.pack_start(warning_label, False, False, 0)

        # Progress area
        progress_frame = Gtk.Frame()
        progress_frame.set_size_request(800, 200)
        progress_frame.set_shadow_type(Gtk.ShadowType.IN)
        progress_frame.set_label("Analysis Progress")
        main_vbox.pack_start(progress_frame, False, False, 0)

        progress_vbox = Gtk.VBox(spacing=10)
        progress_vbox.set_property("margin", 15)
        progress_frame.add(progress_vbox)

        self.progress_info = Gtk.Label()
        self.progress_info.set_markup("<span font='14'>Ready to start analysis</span>")
        progress_vbox.pack_start(self.progress_info, False, False, 0)

        self.current_test_info = Gtk.Label()
        self.current_test_info.set_markup("")
        progress_vbox.pack_start(self.current_test_info, False, False, 0)

        self.progress_bar_label = Gtk.Label()
        self.progress_bar_label.set_markup("")
        progress_vbox.pack_start(self.progress_bar_label, False, False, 0)

        # Device info area
        device_frame = Gtk.Frame()
        device_frame.set_size_request(800, 830)
        device_frame.set_shadow_type(Gtk.ShadowType.IN)
        main_vbox.pack_start(device_frame, True, True, 0)

        device_info_label = Gtk.Label()
        device_info_text = self.generate_device_info_text()
        device_info_label.set_markup(device_info_text)
        device_frame.add(device_info_label)

        self.show_all()

    def generate_device_info_text(self):
        """Generate device information text"""
        if not self.video_devices:
            return "<span font='14'>No video devices found</span>"

        info_text = "<span font='14'><b>Devices to be analyzed:</b>\n\n"

        for device_info in self.video_devices:
            device_path = device_info['path']
            capabilities = device_info['capabilities']

            info_text += f"<b>{device_path}:</b>\n"

            for fmt, fmt_data in capabilities.items():
                combinations = sum(len(fps_list) for fps_list in fmt_data['resolutions'].values())
                info_text += f"  • {fmt}: {len(fmt_data['resolutions'])} resolutions, {combinations} combinations\n"

                for resolution, fps_list in fmt_data['resolutions'].items():
                    w, h = resolution
                    fps_str = ', '.join([f"{fps:.0f}" for fps in sorted(fps_list)])
                    info_text += f"    {w}x{h}: {fps_str} fps\n"

            info_text += "\n"

        info_text += f"<b>Output:</b> {self.output_excel}\n"
        info_text += "<b>Process:</b> Record 15s → Measure file → Delete → Update Excel</span>"

        return info_text

    def start_complete_analysis(self, btn):
        """Start the complete automated analysis"""
        if self.is_analyzing:
            return

        self.is_analyzing = True
        self.completed_combinations = 0

        # Disable button
        self.start_analysis_btn.set_sensitive(False)
        self.start_analysis_btn.set_label("⏳ Analysis Running...")

        self.progress_info.set_markup("<span font='14' color='blue'><b>🚀 Starting complete analysis...</b></span>")

        # Start the analysis
        GLib.timeout_add(1000, self.run_next_test)

    def run_next_test(self):
        """Run the next test in sequence"""
        if not self.is_analyzing:
            return False

        # Find next combination to test
        next_test = self.get_next_test_combination()

        if not next_test:
            # All tests complete
            self.complete_analysis()
            return False

        self.current_test = next_test
        device_path = next_test['device_path']
        format_name = next_test['format']
        resolution = next_test['resolution']
        fps = next_test['fps']

        w, h = resolution

        # Update progress display
        progress_pct = (self.completed_combinations / self.total_combinations) * 100
        progress_bar = "█" * int(progress_pct // 5) + "░" * (20 - int(progress_pct // 5))

        self.progress_info.set_markup(f"<span font='14' color='blue'><b>Progress: {self.completed_combinations}/{self.total_combinations} ({progress_pct:.1f}%)</b></span>")
        self.current_test_info.set_markup(f"<span font='12'>Testing: {device_path} {format_name} {w}x{h}@{fps}fps</span>")
        self.progress_bar_label.set_markup(f"<span font='10'>[{progress_bar}]</span>")

        # Start recording this combination
        self.record_test_video()

        return False  # Don't continue timer - finish_test_recording will schedule next

    def get_next_test_combination(self):
        """Get the next combination to test"""
        current_index = 0

        for device_info in self.video_devices:
            device_path = device_info['path']
            capabilities = device_info['capabilities']

            for format_name, format_data in capabilities.items():
                for resolution, fps_list in format_data['resolutions'].items():
                    for fps in sorted(fps_list):
                        if current_index == self.completed_combinations:
                            return {
                                'device_path': device_path,
                                'format': format_name,
                                'resolution': resolution,
                                'fps': fps
                            }
                        current_index += 1

        return None  # No more combinations

    def record_test_video(self):
        """Record a test video for the current combination"""
        try:
            device_path = self.current_test['device_path']
            format_name = self.current_test['format']
            resolution = self.current_test['resolution']
            fps = self.current_test['fps']

            w, h = resolution

            # Generate temp filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if format_name == 'H264':
                ext = 'mp4'
            else:
                ext = 'avi'

            filename = f"test_{format_name}_{w}x{h}_{fps:.0f}fps_{timestamp}.{ext}"
            output_file = os.path.join(self.temp_dir, filename)

            # Build recording pipeline
            if format_name == 'H264':
                caps = f"video/x-h264,width={w},height={h},framerate={fps:.0f}/1"
                pipeline_str = f"v4l2src device={device_path} ! {caps} ! h264parse ! mp4mux ! filesink location={output_file}"

            elif format_name == 'MJPG':
                caps = f"image/jpeg,width={w},height={h},framerate={fps:.0f}/1"
                pipeline_str = f"v4l2src device={device_path} ! {caps} ! avimux ! filesink location={output_file}"

            else:  # YUYV
                caps = f"video/x-raw,format=YUY2,width={w},height={h},framerate={fps:.0f}/1"
                pipeline_str = f"v4l2src device={device_path} ! {caps} ! videoconvert ! x264enc ! avimux ! filesink location={output_file}"

            print(f"Testing: {pipeline_str}")

            # Try to create and start pipeline
            self.pipeline = Gst.parse_launch(pipeline_str)

            # Simple approach: start pipeline and wait exactly 16 seconds
            ret = self.pipeline.set_state(Gst.State.PLAYING)

            self.is_recording = True
            self.current_test['output_file'] = output_file
            self.current_test['record_start'] = time.time()

            # Wait exactly 16 seconds then finish
            GLib.timeout_add(self.wait_duration * 1000, self.finish_test_recording)

        except Exception as e:
            print(f"Recording error: {e}")
            # Simple cleanup and mark as failed
            if self.pipeline:
                try:
                    self.pipeline.set_state(Gst.State.NULL)
                    self.pipeline = None
                except:
                    pass

            self.record_test_result(False, 0, 0)
            self.completed_combinations += 1
            GLib.timeout_add(100, self.run_next_test)


    def finish_test_recording(self):
        """Finish recording and measure results - simple approach"""
        try:
            if self.pipeline:
                # Send EOS to properly finish the file
                self.pipeline.send_event(Gst.Event.new_eos())
                time.sleep(0.5)  # Wait for EOS to process
                self.pipeline.set_state(Gst.State.NULL)
                self.pipeline = None

            self.is_recording = False

            output_file = self.current_test['output_file']

            # Check if file was created and measure it
            if os.path.exists(output_file):
                file_size = os.path.getsize(output_file)
                file_size_mb = file_size / (1024 * 1024)

                # Simple check: if file exists with any reasonable size, consider success
                if file_size > 1024:  # At least 1KB
                    # Calculate actual bitrate based on 15-second recording
                    bits = file_size * 8
                    bitrate_bps = bits / self.recording_duration
                    bitrate_kbps = bitrate_bps / 1000

                    print(f"Success: {output_file} - {file_size_mb:.2f} MB, {bitrate_kbps:.1f} kbps")
                    self.record_test_result(True, file_size_mb, bitrate_kbps)
                else:
                    print(f"Failed: {output_file} too small ({file_size} bytes)")
                    self.record_test_result(False, 0, 0)

                # Always delete the file to save space
                try:
                    os.remove(output_file)
                except:
                    pass

            else:
                print(f"Failed: {output_file} not created")
                self.record_test_result(False, 0, 0)

        except Exception as e:
            print(f"Error finishing recording: {e}")
            self.record_test_result(False, 0, 0)

        self.completed_combinations += 1

        # Schedule next test
        GLib.timeout_add(500, self.run_next_test)

        return False  # Don't repeat this timer

    def record_test_result(self, success, file_size_mb, bitrate_kbps):
        """Record the result of a test"""
        device_path = self.current_test['device_path']
        format_name = self.current_test['format']
        resolution = self.current_test['resolution']
        fps = self.current_test['fps']

        # Store result
        if device_path not in self.analysis_results:
            self.analysis_results[device_path] = {}

        if format_name not in self.analysis_results[device_path]:
            self.analysis_results[device_path][format_name] = []

        self.analysis_results[device_path][format_name].append({
            'resolution': resolution,
            'fps': fps,
            'success': success,
            'file_size_mb': file_size_mb,
            'bitrate_kbps': bitrate_kbps
        })

    def complete_analysis(self):
        """Complete the analysis and generate Excel file"""
        self.is_analyzing = False

        self.progress_info.set_markup("<span font='14' color='green'><b>✅ Analysis Complete! Generating Excel file...</b></span>")
        self.current_test_info.set_markup("")
        self.progress_bar_label.set_markup("")

        # Generate Excel file
        try:
            self.generate_real_excel_file()
            self.progress_info.set_markup(f"<span font='14' color='green'><b>🎉 Complete! Excel file saved: {self.output_excel}</b></span>")
        except Exception as e:
            self.progress_info.set_markup(f"<span font='14' color='red'><b>❌ Excel generation failed: {e}</b></span>")

        # Re-enable button
        self.start_analysis_btn.set_sensitive(True)
        self.start_analysis_btn.set_label("🚀 Start Complete Analysis")

        # Cleanup temp directory
        try:
            os.rmdir(self.temp_dir)
        except:
            pass

    def generate_real_excel_file(self):
        """Generate Excel file with real measured data"""
        print(f"Generating Excel file: {self.output_excel}")

        wb = Workbook()
        wb.remove(wb.active)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Process each device/format combination
        for device_path, device_data in self.analysis_results.items():
            device_name = device_path.replace('/dev/', '')

            for format_name, format_results in device_data.items():
                if not format_results:
                    continue

                # Create worksheet
                sheet_name = f"{device_name}_{format_name}"
                ws = wb.create_sheet(title=sheet_name)

                # Convert to DataFrame
                df_data = []
                for result in format_results:
                    w, h = result['resolution']
                    df_data.append({
                        'Resolution': f"{w}x{h}",
                        'Width': w,
                        'Height': h,
                        'FPS': result['fps'],
                        'Real Bitrate (kbps)': round(result['bitrate_kbps'], 1) if result['success'] else 0,
                        'Real File Size 15s (MB)': round(result['file_size_mb'], 3) if result['success'] else 0,
                        'Works': "✓" if result['success'] else "✗"
                    })

                df = pd.DataFrame(df_data)

                # Write title
                ws['A1'] = f"REAL DATA: {device_path} - {format_name}"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:H1')

                # Create matrix
                resolutions = df['Resolution'].unique()
                fps_values = sorted(df['FPS'].unique())

                # Headers
                row = 3
                ws[f'A{row}'] = "Resolution"
                ws[f'A{row}'].font = header_font
                ws[f'A{row}'].fill = header_fill
                ws[f'A{row}'].alignment = center_align
                ws[f'A{row}'].border = border

                col = 2
                for fps in fps_values:
                    cell = ws.cell(row=row, column=col, value=f"{fps} FPS")
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border
                    col += 1

                # Fill matrix
                row += 1
                for resolution in resolutions:
                    ws.cell(row=row, column=1, value=resolution).font = Font(bold=True)
                    ws.cell(row=row, column=1).alignment = center_align
                    ws.cell(row=row, column=1).border = border

                    col = 2
                    for fps in fps_values:
                        matching = df[(df['Resolution'] == resolution) & (df['FPS'] == fps)]

                        if not matching.empty:
                            data = matching.iloc[0]
                            bitrate = data['Real Bitrate (kbps)']
                            filesize = data['Real File Size 15s (MB)']
                            works = data['Works']

                            if works == "✓":
                                cell_content = f"{bitrate} kbps\n{filesize} MB\n✓ REAL"
                                cell_fill = success_fill
                            else:
                                cell_content = "FAILED\n0 MB\n✗"
                                cell_fill = fail_fill

                            cell = ws.cell(row=row, column=col, value=cell_content)
                            cell.alignment = center_align
                            cell.border = border
                            cell.fill = cell_fill

                        else:
                            cell = ws.cell(row=row, column=col, value="N/A")
                            cell.alignment = center_align
                            cell.border = border

                        col += 1
                    row += 1

                # Add detailed table
                row += 2
                ws.cell(row=row, column=1, value="REAL MEASURED DATA:").font = Font(bold=True, size=12)
                row += 1

                headers = ['Resolution', 'FPS', 'Real Bitrate (kbps)', 'Real File Size 15s (MB)', 'Works']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border

                row += 1
                for _, data in df.iterrows():
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=data[header])
                        cell.alignment = center_align
                        cell.border = border

                        if header == 'Works':
                            if data[header] == "✓":
                                cell.fill = success_fill
                            else:
                                cell.fill = fail_fill
                    row += 1

                # Auto-adjust column widths
                for col_num in range(1, ws.max_column + 1):
                    max_length = 0
                    column_letter = get_column_letter(col_num)

                    for row_num in range(1, ws.max_row + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 20)
                    ws.column_dimensions[column_letter].width = adjusted_width

        # Create summary sheet
        summary_ws = wb.create_sheet(title="REAL_DATA_Summary", index=0)
        summary_ws['A1'] = "REAL Camera Analysis Summary"
        summary_ws['A1'].font = Font(bold=True, size=16)

        row = 3
        total_tested = 0
        total_successful = 0

        for device_path, device_data in self.analysis_results.items():
            summary_ws.cell(row=row, column=1, value=f"Device: {device_path}").font = Font(bold=True)
            row += 1

            for format_name, format_results in device_data.items():
                successful = len([r for r in format_results if r['success']])
                total = len(format_results)

                total_tested += total
                total_successful += successful

                summary_ws.cell(row=row, column=2, value=f"{format_name}: {successful}/{total} combinations successful")
                row += 1
            row += 1

        summary_ws.cell(row=row, column=1, value=f"TOTAL: {total_successful}/{total_tested} combinations successful").font = Font(bold=True, size=14)

        wb.save(self.output_excel)
        print(f"Real analysis Excel file saved: {self.output_excel}")

if __name__ == "__main__":
    app = RealCameraAnalyzer()
    Gtk.main()