#!/usr/bin/env python3
# Console Camera Analysis - Lightweight terminal-based version for embedded systems

import os
import sys
import glob
import subprocess
import time
import re
import threading
from datetime import datetime

# Check if required modules are available
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Required module missing: {e}")
    print("Please install required modules:")
    print("pip install pandas openpyxl")
    sys.exit(1)

try:
    import gi
    gi.require_version("Gst", "1.0")
    from gi.repository import Gst, GLib
except ImportError as e:
    print(f"Required module missing: {e}")
    print("GStreamer python bindings required")
    sys.exit(1)

class ConsoleCameraAnalyzer:
    def __init__(self):
        print("🎥 Console Camera Analysis Tool")
        print("=" * 50)

        # Initialize GStreamer
        Gst.init(None)

        # Analysis settings
        self.recording_duration = 15  # seconds
        self.wait_duration = 16  # seconds
        self.temp_dir = "temp_analysis"
        self.output_excel = "console_camera_analysis.xlsx"

        # Analysis state
        self.video_devices = []
        self.analysis_results = {}
        self.current_test = {}
        self.total_combinations = 0
        self.completed_combinations = 0
        self.is_analyzing = False
        self.is_recording = False
        self.pipeline = None
        self.main_loop = None

        # Find video devices and their capabilities
        print("\n🔍 Scanning video devices for capabilities...")
        self.get_real_device_capabilities()
        self.create_temp_directory()

        if self.total_combinations == 0:
            print("❌ No video devices found with usable capabilities!")
            sys.exit(1)

        print(f"\n✅ Ready to analyze {self.total_combinations} combinations")
        print(f"📁 Temporary files: {self.temp_dir}")
        print(f"📊 Output Excel: {self.output_excel}")

    def parse_v4l2_output(self, device_path):
        """Parse v4l2-ctl output to extract real device capabilities"""
        try:
            result = subprocess.run(['v4l2-ctl', '--device', device_path, '--list-formats-ext'],
                                  capture_output=True, text=True, timeout=5)

            if result.returncode != 0 or not result.stdout.strip():
                return {}

            capabilities = {}
            current_format = None

            lines = result.stdout.split('\n')

            for line in lines:
                line = line.strip()

                # Look for format lines
                format_match = re.search(r"\[(\d+)\]:\s+'([^']+)'\s+\(([^)]+)\)", line)
                if format_match:
                    format_code = format_match.group(2)
                    format_desc = format_match.group(3)
                    current_format = format_code
                    capabilities[current_format] = {
                        'description': format_desc,
                        'resolutions': {}
                    }
                    continue

                # Look for size lines
                size_match = re.search(r"Size:\s+Discrete\s+(\d+)x(\d+)", line)
                if size_match and current_format:
                    width = int(size_match.group(1))
                    height = int(size_match.group(2))
                    resolution = (width, height)

                    if resolution not in capabilities[current_format]['resolutions']:
                        capabilities[current_format]['resolutions'][resolution] = []
                    continue

                # Look for interval lines
                interval_match = re.search(r"Interval:\s+Discrete\s+[\d.]+s\s+\(([\d.]+)\s+fps\)", line)
                if interval_match and current_format:
                    fps = float(interval_match.group(1))
                    # Add this fps to the last resolution found
                    resolutions = capabilities[current_format]['resolutions']
                    if resolutions:
                        last_resolution = list(resolutions.keys())[-1]
                        capabilities[current_format]['resolutions'][last_resolution].append(fps)

            return capabilities

        except Exception as e:
            print(f"⚠️  Error parsing v4l2 output for {device_path}: {e}")
            return {}

    def get_real_device_capabilities(self):
        """Get video devices and their REAL capabilities from v4l2-ctl"""
        for device_path in glob.glob('/dev/video*'):
            try:
                print(f"   Checking {device_path}...")
                capabilities = self.parse_v4l2_output(device_path)

                if capabilities:
                    device_info = {
                        'path': device_path,
                        'capabilities': capabilities
                    }
                    self.video_devices.append(device_info)

                    # Count total combinations
                    for format_name, format_data in capabilities.items():
                        for resolution, fps_list in format_data['resolutions'].items():
                            self.total_combinations += len(fps_list)

                    print(f"   ✅ {device_path}: {len(capabilities)} formats")
                else:
                    print(f"   ❌ {device_path}: No usable formats")

            except Exception as e:
                print(f"   ⚠️  Error checking {device_path}: {e}")
                continue

        print(f"\n📊 Found {len(self.video_devices)} usable video devices")
        print(f"🎯 Total combinations to test: {self.total_combinations}")

    def create_temp_directory(self):
        """Create temporary directory for test recordings"""
        try:
            if not os.path.exists(self.temp_dir):
                os.makedirs(self.temp_dir)
                print(f"📁 Created temporary directory: {self.temp_dir}")
        except Exception as e:
            print(f"❌ Error creating temp directory: {e}")

    def show_device_summary(self):
        """Show a summary of detected devices"""
        print("\n" + "=" * 50)
        print("📋 DEVICE SUMMARY")
        print("=" * 50)

        for device_info in self.video_devices:
            device_path = device_info['path']
            capabilities = device_info['capabilities']

            print(f"\n🎥 Device: {device_path}")

            for fmt, fmt_data in capabilities.items():
                combinations = sum(len(fps_list) for fps_list in fmt_data['resolutions'].values())
                print(f"   📺 {fmt}: {len(fmt_data['resolutions'])} resolutions, {combinations} combinations")

                for resolution, fps_list in fmt_data['resolutions'].items():
                    w, h = resolution
                    fps_str = ', '.join([f"{fps:.0f}" for fps in sorted(fps_list)])
                    print(f"      {w}x{h}: {fps_str} fps")

    def start_complete_analysis(self):
        """Start the complete automated analysis"""
        print("\n" + "=" * 50)
        print("🚀 STARTING COMPLETE ANALYSIS")
        print("=" * 50)

        estimated_time = self.total_combinations * self.wait_duration
        print(f"⏱️  Estimated time: {estimated_time} seconds ({estimated_time // 60} minutes)")
        print(f"📝 Process: Record 15s → Wait 16s → Measure → Delete → Repeat")
        print()

        # Confirm start
        try:
            response = input("🤔 Do you want to start the analysis? (y/N): ").strip().lower()
            if response not in ['y', 'yes']:
                print("❌ Analysis cancelled by user")
                return False
        except KeyboardInterrupt:
            print("\n❌ Analysis cancelled by user")
            return False

        self.is_analyzing = True
        self.completed_combinations = 0
        self.analysis_results = {}

        # Start GLib main loop in a separate thread
        self.main_loop = GLib.MainLoop()
        loop_thread = threading.Thread(target=self.main_loop.run)
        loop_thread.daemon = True
        loop_thread.start()

        # Start the first test
        GLib.timeout_add(100, self.run_next_test)

        # Keep main thread alive and show progress
        try:
            while self.is_analyzing:
                time.sleep(1)
                self.show_progress()
        except KeyboardInterrupt:
            print("\n🛑 Analysis interrupted by user")
            self.is_analyzing = False
            if self.main_loop:
                self.main_loop.quit()
            return False

        return True

    def show_progress(self):
        """Show analysis progress"""
        if not self.is_analyzing:
            return

        progress_pct = (self.completed_combinations / self.total_combinations) * 100 if self.total_combinations > 0 else 0
        progress_bar_length = 40
        filled_length = int(progress_bar_length * progress_pct / 100)

        bar = "█" * filled_length + "░" * (progress_bar_length - filled_length)

        # Show progress and current test on same line
        current_test_info = ""
        if hasattr(self, 'current_test') and self.current_test and self.is_recording:
            device_path = self.current_test.get('device_path', '')
            format_name = self.current_test.get('format', '')
            resolution = self.current_test.get('resolution', (0, 0))
            fps = self.current_test.get('fps', 0)

            w, h = resolution
            current_test_info = f" | 🎬 {device_path} {format_name} {w}x{h}@{fps}fps"

        # Clear line and show progress with current test info
        line = f"\r📊 [{bar}] {self.completed_combinations}/{self.total_combinations} ({progress_pct:.1f}%){current_test_info}"
        # Pad with spaces to clear any leftover characters
        line = line.ljust(120)
        print(line, end="", flush=True)

    def run_next_test(self):
        """Run the next test in sequence"""
        if not self.is_analyzing:
            return False

        # Find next combination to test
        next_test = self.get_next_test_combination()

        if not next_test:
            # All tests complete
            self.complete_analysis()
            return False

        self.current_test = next_test
        device_path = next_test['device_path']
        format_name = next_test['format']
        resolution = next_test['resolution']
        fps = next_test['fps']

        w, h = resolution
        print(f"\n🎬 Testing: {device_path} {format_name} {w}x{h}@{fps}fps", flush=True)

        # Start recording this combination
        self.record_test_video()

        return False  # Don't repeat this timer

    def get_next_test_combination(self):
        """Get the next combination to test"""
        current_index = 0

        for device_info in self.video_devices:
            device_path = device_info['path']
            capabilities = device_info['capabilities']

            for format_name, format_data in capabilities.items():
                for resolution, fps_list in format_data['resolutions'].items():
                    for fps in sorted(fps_list):
                        if current_index == self.completed_combinations:
                            return {
                                'device_path': device_path,
                                'format': format_name,
                                'resolution': resolution,
                                'fps': fps
                            }
                        current_index += 1

        return None  # No more combinations

    def record_test_video(self):
        """Record a test video for the current combination"""
        try:
            device_path = self.current_test['device_path']
            format_name = self.current_test['format']
            resolution = self.current_test['resolution']
            fps = self.current_test['fps']

            w, h = resolution

            # Generate temp filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if format_name == 'H264':
                ext = 'mp4'
            else:
                ext = 'avi'

            filename = f"test_{format_name}_{w}x{h}_{fps:.0f}fps_{timestamp}.{ext}"
            output_file = os.path.join(self.temp_dir, filename)

            # Build recording pipeline
            if format_name == 'H264':
                caps = f"video/x-h264,width={w},height={h},framerate={fps:.0f}/1"
                pipeline_str = f"v4l2src device={device_path} ! {caps} ! h264parse ! mp4mux ! filesink location={output_file}"

            elif format_name == 'MJPG':
                caps = f"image/jpeg,width={w},height={h},framerate={fps:.0f}/1"
                pipeline_str = f"v4l2src device={device_path} ! {caps} ! avimux ! filesink location={output_file}"

            else:  # YUYV
                caps = f"video/x-raw,format=YUY2,width={w},height={h},framerate={fps:.0f}/1"
                pipeline_str = f"v4l2src device={device_path} ! {caps} ! videoconvert ! x264enc ! avimux ! filesink location={output_file}"

            # Try to create and start pipeline
            self.pipeline = Gst.parse_launch(pipeline_str)
            ret = self.pipeline.set_state(Gst.State.PLAYING)

            self.is_recording = True
            self.current_test['output_file'] = output_file
            self.current_test['record_start'] = time.time()

            # Schedule finish after wait_duration
            GLib.timeout_add(self.wait_duration * 1000, self.finish_test_recording)

        except Exception as e:
            print(f"❌ Recording error: {e}")
            # Simple cleanup and mark as failed
            if self.pipeline:
                try:
                    self.pipeline.set_state(Gst.State.NULL)
                    self.pipeline = None
                except:
                    pass

            self.record_test_result(False, 0, 0)
            self.completed_combinations += 1
            # Schedule next test
            GLib.timeout_add(100, self.run_next_test)

    def finish_test_recording(self):
        """Finish recording and measure results - simple approach"""
        try:
            if self.pipeline:
                # Send EOS to properly finish the file
                self.pipeline.send_event(Gst.Event.new_eos())
                time.sleep(0.5)  # Wait for EOS to process
                self.pipeline.set_state(Gst.State.NULL)
                self.pipeline = None

            self.is_recording = False

            output_file = self.current_test['output_file']

            # Check if file was created and measure it
            if os.path.exists(output_file):
                file_size = os.path.getsize(output_file)
                file_size_mb = file_size / (1024 * 1024)

                # Simple check: if file exists with any reasonable size, consider success
                if file_size > 1024:  # At least 1KB
                    # Calculate actual bitrate based on 15-second recording
                    bits = file_size * 8
                    bitrate_bps = bits / self.recording_duration
                    bitrate_kbps = bitrate_bps / 1000

                    print(f"\n   ✅ Success: {file_size_mb:.2f} MB, {bitrate_kbps:.1f} kbps", flush=True)
                    self.record_test_result(True, file_size_mb, bitrate_kbps)
                else:
                    print(f"\n   ❌ Failed: File too small ({file_size} bytes)", flush=True)
                    self.record_test_result(False, 0, 0)

                # Always delete the file to save space
                try:
                    os.remove(output_file)
                except:
                    pass

            else:
                print(f"\n   ❌ Failed: File not created", flush=True)
                self.record_test_result(False, 0, 0)

        except Exception as e:
            print(f"❌ Error finishing recording: {e}")
            self.record_test_result(False, 0, 0)

        self.completed_combinations += 1

        # Schedule next test
        GLib.timeout_add(500, self.run_next_test)

        return False  # Don't repeat this timer

    def record_test_result(self, success, file_size_mb, bitrate_kbps):
        """Record the result of a test"""
        device_path = self.current_test['device_path']
        format_name = self.current_test['format']
        resolution = self.current_test['resolution']
        fps = self.current_test['fps']

        # Store result
        if device_path not in self.analysis_results:
            self.analysis_results[device_path] = {}

        if format_name not in self.analysis_results[device_path]:
            self.analysis_results[device_path][format_name] = []

        self.analysis_results[device_path][format_name].append({
            'resolution': resolution,
            'fps': fps,
            'success': success,
            'file_size_mb': file_size_mb,
            'bitrate_kbps': bitrate_kbps
        })

    def complete_analysis(self):
        """Complete the analysis and generate Excel file"""
        self.is_analyzing = False
        print(f"\n\n🎉 Analysis complete! Generating Excel file...")

        # Generate Excel file
        try:
            self.generate_real_excel_file()
            print(f"✅ Excel file saved: {self.output_excel}")
        except Exception as e:
            print(f"❌ Excel generation failed: {e}")

        # Show summary
        self.show_results_summary()

        # Cleanup temp directory
        try:
            os.rmdir(self.temp_dir)
            print(f"🗑️  Cleaned up temporary directory")
        except:
            pass

        if self.main_loop:
            self.main_loop.quit()

    def show_results_summary(self):
        """Show a summary of results"""
        print("\n" + "=" * 50)
        print("📊 ANALYSIS RESULTS SUMMARY")
        print("=" * 50)

        total_tested = 0
        total_successful = 0

        for device_path, device_data in self.analysis_results.items():
            print(f"\n🎥 Device: {device_path}")

            for format_name, format_results in device_data.items():
                successful = len([r for r in format_results if r['success']])
                total = len(format_results)

                total_tested += total
                total_successful += successful

                status = "✅" if successful > 0 else "❌"
                print(f"   {status} {format_name}: {successful}/{total} combinations successful")

                if successful > 0:
                    # Show best and worst performing combinations
                    successful_results = [r for r in format_results if r['success']]
                    best = max(successful_results, key=lambda x: x['bitrate_kbps'])
                    worst = min(successful_results, key=lambda x: x['bitrate_kbps'])

                    w_best, h_best = best['resolution']
                    w_worst, h_worst = worst['resolution']

                    print(f"      🏆 Best: {w_best}x{h_best}@{best['fps']}fps - {best['bitrate_kbps']:.1f} kbps")
                    print(f"      📉 Worst: {w_worst}x{h_worst}@{worst['fps']}fps - {worst['bitrate_kbps']:.1f} kbps")

        print(f"\n🎯 OVERALL: {total_successful}/{total_tested} combinations successful ({(total_successful/total_tested*100):.1f}%)")

    def generate_real_excel_file(self):
        """Generate Excel file with real measured data"""
        print(f"📊 Generating Excel file: {self.output_excel}")

        wb = Workbook()
        wb.remove(wb.active)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Process each device/format combination
        for device_path, device_data in self.analysis_results.items():
            device_name = device_path.replace('/dev/', '')

            for format_name, format_results in device_data.items():
                if not format_results:
                    continue

                # Create worksheet
                sheet_name = f"{device_name}_{format_name}"
                ws = wb.create_sheet(title=sheet_name)

                # Convert to DataFrame
                df_data = []
                for result in format_results:
                    w, h = result['resolution']
                    df_data.append({
                        'Resolution': f"{w}x{h}",
                        'Width': w,
                        'Height': h,
                        'FPS': result['fps'],
                        'Real Bitrate (kbps)': round(result['bitrate_kbps'], 1) if result['success'] else 0,
                        'Real File Size 15s (MB)': round(result['file_size_mb'], 3) if result['success'] else 0,
                        'Works': "✓" if result['success'] else "✗"
                    })

                df = pd.DataFrame(df_data)

                # Write title
                ws['A1'] = f"CONSOLE REAL DATA: {device_path} - {format_name}"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:H1')

                # Data table
                row = 3
                headers = ['Resolution', 'FPS', 'Real Bitrate (kbps)', 'Real File Size 15s (MB)', 'Works']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border

                row += 1
                for _, data in df.iterrows():
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=data[header])
                        cell.alignment = center_align
                        cell.border = border

                        if header == 'Works':
                            if data[header] == "✓":
                                cell.fill = success_fill
                            else:
                                cell.fill = fail_fill
                    row += 1

        # Create summary sheet
        summary_ws = wb.create_sheet(title="CONSOLE_Summary", index=0)
        summary_ws['A1'] = "Console Real Camera Analysis Summary"
        summary_ws['A1'].font = Font(bold=True, size=16)

        row = 3
        total_tested = 0
        total_successful = 0

        for device_path, device_data in self.analysis_results.items():
            summary_ws.cell(row=row, column=1, value=f"Device: {device_path}").font = Font(bold=True)
            row += 1

            for format_name, format_results in device_data.items():
                successful = len([r for r in format_results if r['success']])
                total = len(format_results)

                total_tested += total
                total_successful += successful

                summary_ws.cell(row=row, column=2, value=f"{format_name}: {successful}/{total} combinations successful")
                row += 1
            row += 1

        summary_ws.cell(row=row, column=1, value=f"TOTAL: {total_successful}/{total_tested} combinations successful").font = Font(bold=True, size=14)

        wb.save(self.output_excel)

    def run_interactive(self):
        """Run interactive console interface"""
        while True:
            print("\n" + "=" * 50)
            print("🎮 INTERACTIVE MENU")
            print("=" * 50)
            print("1. 📋 Show device summary")
            print("2. 🚀 Start complete analysis")
            print("3. 🚪 Exit")

            try:
                choice = input("\n🤔 Choose an option (1-3): ").strip()

                if choice == '1':
                    self.show_device_summary()
                elif choice == '2':
                    success = self.start_complete_analysis()
                    if success:
                        break  # Exit after successful analysis
                elif choice == '3':
                    print("👋 Goodbye!")
                    break
                else:
                    print("❌ Invalid choice. Please enter 1, 2, or 3.")

            except KeyboardInterrupt:
                print("\n👋 Goodbye!")
                break

if __name__ == "__main__":
    try:
        analyzer = ConsoleCameraAnalyzer()
        analyzer.run_interactive()
    except KeyboardInterrupt:
        print("\n👋 Application interrupted by user")
    except Exception as e:
        print(f"❌ Application error: {e}")
        import traceback
        traceback.print_exc()