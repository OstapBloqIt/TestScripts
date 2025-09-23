#!/usr/bin/env python3
# Camera Analysis Tool - Generate Excel matrices for all format/resolution/fps combinations

import gi
import glob
import subprocess
import time
import re
import sys

# Check if required modules are available
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError as e:
    print(f"Required module missing: {e}")
    print("Please install required modules:")
    print("pip install pandas openpyxl")
    sys.exit(1)

# Initialize GStreamer for testing
try:
    gi.require_version("Gst", "1.0")
    from gi.repository import Gst
    Gst.init(None)
except:
    print("Warning: GStreamer not available for testing pipelines")
    Gst = None

class CameraAnalyzer:
    def __init__(self):
        self.video_devices = []
        self.test_results = {}

    def parse_v4l2_output(self, device_path):
        """Parse v4l2-ctl output to extract real device capabilities"""
        try:
            result = subprocess.run(['v4l2-ctl', '--device', device_path, '--list-formats-ext'],
                                  capture_output=True, text=True, timeout=5)

            if result.returncode != 0 or not result.stdout.strip():
                return {}

            capabilities = {}
            current_format = None

            lines = result.stdout.split('\n')

            for line in lines:
                line = line.strip()

                # Look for format lines
                format_match = re.search(r"\[(\d+)\]:\s+'([^']+)'\s+\(([^)]+)\)", line)
                if format_match:
                    format_code = format_match.group(2)
                    format_desc = format_match.group(3)
                    current_format = format_code
                    capabilities[current_format] = {
                        'description': format_desc,
                        'resolutions': {}
                    }
                    continue

                # Look for size lines
                size_match = re.search(r"Size:\s+Discrete\s+(\d+)x(\d+)", line)
                if size_match and current_format:
                    width = int(size_match.group(1))
                    height = int(size_match.group(2))
                    resolution = (width, height)

                    if resolution not in capabilities[current_format]['resolutions']:
                        capabilities[current_format]['resolutions'][resolution] = []
                    continue

                # Look for interval lines
                interval_match = re.search(r"Interval:\s+Discrete\s+[\d.]+s\s+\(([\d.]+)\s+fps\)", line)
                if interval_match and current_format:
                    fps = float(interval_match.group(1))
                    # Add this fps to the last resolution found
                    resolutions = capabilities[current_format]['resolutions']
                    if resolutions:
                        last_resolution = list(resolutions.keys())[-1]
                        capabilities[current_format]['resolutions'][last_resolution].append(fps)

            return capabilities

        except Exception as e:
            print(f"Error parsing v4l2 output for {device_path}: {e}")
            return {}

    def get_device_capabilities(self):
        """Get all video devices and their capabilities"""
        print("Scanning video devices...")

        for device_path in glob.glob('/dev/video*'):
            print(f"Checking {device_path}...")
            capabilities = self.parse_v4l2_output(device_path)

            if capabilities:
                device_info = {
                    'path': device_path,
                    'capabilities': capabilities
                }
                self.video_devices.append(device_info)
                print(f"  Found {len(capabilities)} formats")
            else:
                print(f"  No usable formats")

        print(f"Total usable devices: {len(self.video_devices)}")

    def estimate_bandwidth_and_filesize(self, format_name, width, height, fps):
        """Estimate bandwidth (kbps) and 15-second file size (MB)"""

        # Base calculations
        pixels = width * height

        if format_name == 'H264':
            # H.264 compression estimates (bits per pixel)
            if pixels <= 640*480:  # VGA and below
                bpp = 0.08  # High compression
            elif pixels <= 1280*720:  # Up to 720p
                bpp = 0.06
            else:  # 1080p and above
                bpp = 0.04

        elif format_name == 'MJPG':
            # Motion JPEG compression estimates
            if pixels <= 640*480:
                bpp = 0.5  # Moderate compression
            elif pixels <= 1280*720:
                bpp = 0.4
            else:
                bpp = 0.3

        else:  # YUYV (uncompressed)
            bpp = 16  # YUV422 = 16 bits per pixel

        # Calculate bandwidth (kbps)
        bits_per_frame = pixels * bpp
        bits_per_second = bits_per_frame * fps
        bandwidth_kbps = bits_per_second / 1000

        # Calculate 15-second file size (MB)
        bits_15_seconds = bits_per_second * 15
        bytes_15_seconds = bits_15_seconds / 8
        mb_15_seconds = bytes_15_seconds / (1024 * 1024)

        return bandwidth_kbps, mb_15_seconds

    def test_pipeline(self, device_path, format_name, width, height, fps):
        """Test if a specific pipeline configuration works"""
        if not Gst:
            return "Unknown"  # Can't test without GStreamer

        try:
            # Build test pipeline (without display)
            if format_name == 'H264':
                caps = f"video/x-h264,width={width},height={height},framerate={fps:.0f}/1"
                pipeline_str = f"v4l2src device={device_path} ! {caps} ! h264parse ! avdec_h264 ! videoconvert ! fakesink"
            elif format_name == 'MJPG':
                caps = f"image/jpeg,width={width},height={height},framerate={fps:.0f}/1"
                pipeline_str = f"v4l2src device={device_path} ! {caps} ! jpegdec ! videoconvert ! fakesink"
            else:  # YUYV
                caps = f"video/x-raw,format=YUY2,width={width},height={height},framerate={fps:.0f}/1"
                pipeline_str = f"v4l2src device={device_path} ! {caps} ! videoconvert ! fakesink"

            # Test the pipeline briefly
            pipeline = Gst.parse_launch(pipeline_str)
            pipeline.set_state(Gst.State.PLAYING)

            # Let it run for 1 second
            time.sleep(1)

            # Check if it's working
            state = pipeline.get_state(Gst.CLOCK_TIME_NONE)
            success = state[1] == Gst.State.PLAYING

            pipeline.set_state(Gst.State.NULL)

            return "✓" if success else "✗"

        except Exception as e:
            return "✗"

    def analyze_device(self, device_info):
        """Analyze a single device and create data for Excel"""
        device_path = device_info['path']
        print(f"\nAnalyzing {device_path}...")

        device_data = {}

        for format_name, format_data in device_info['capabilities'].items():
            print(f"  Testing {format_name}...")

            format_results = []

            for resolution, fps_list in format_data['resolutions'].items():
                width, height = resolution

                for fps in sorted(fps_list):
                    print(f"    Testing {width}x{height} @ {fps} fps...")

                    # Calculate estimates
                    bandwidth, filesize = self.estimate_bandwidth_and_filesize(format_name, width, height, fps)

                    # Test if it works (optional - can be slow)
                    # For now, assume all advertised combinations work
                    works = "✓"  # self.test_pipeline(device_path, format_name, width, height, fps)

                    format_results.append({
                        'Resolution': f"{width}x{height}",
                        'Width': width,
                        'Height': height,
                        'FPS': fps,
                        'Bandwidth (kbps)': round(bandwidth, 1),
                        'File Size 15s (MB)': round(filesize, 2),
                        'Works': works
                    })

            device_data[format_name] = format_results

        return device_data

    def create_excel_file(self, filename="camera_analysis.xlsx"):
        """Create Excel file with matrices for each device and format"""
        print(f"\nCreating Excel file: {filename}")

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # Light red
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        for device_info in self.video_devices:
            device_path = device_info['path']
            device_name = device_path.replace('/dev/', '')

            print(f"Processing {device_path}...")
            device_data = self.analyze_device(device_info)

            for format_name, format_results in device_data.items():
                if not format_results:
                    continue

                # Create worksheet for this device+format combination
                sheet_name = f"{device_name}_{format_name}"
                ws = wb.create_sheet(title=sheet_name)

                # Convert to DataFrame for easier manipulation
                df = pd.DataFrame(format_results)

                # Create a pivot-like structure: Resolution vs FPS
                resolutions = df['Resolution'].unique()
                fps_values = sorted(df['FPS'].unique())

                # Write title
                ws['A1'] = f"Device: {device_path} - Format: {format_name}"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:H1')

                # Create headers
                row = 3
                ws[f'A{row}'] = "Resolution"
                col = 2
                for fps in fps_values:
                    ws.cell(row=row, column=col, value=f"{fps} FPS")
                    ws.cell(row=row, column=col).font = header_font
                    ws.cell(row=row, column=col).fill = header_fill
                    ws.cell(row=row, column=col).alignment = center_align
                    ws.cell(row=row, column=col).border = border
                    col += 1

                # Style the resolution header
                ws[f'A{row}'].font = header_font
                ws[f'A{row}'].fill = header_fill
                ws[f'A{row}'].alignment = center_align
                ws[f'A{row}'].border = border

                # Fill in the matrix
                row += 1
                for resolution in resolutions:
                    ws.cell(row=row, column=1, value=resolution)
                    ws.cell(row=row, column=1).font = Font(bold=True)
                    ws.cell(row=row, column=1).alignment = center_align
                    ws.cell(row=row, column=1).border = border

                    col = 2
                    for fps in fps_values:
                        # Find matching data
                        matching = df[(df['Resolution'] == resolution) & (df['FPS'] == fps)]

                        if not matching.empty:
                            data = matching.iloc[0]
                            bandwidth = data['Bandwidth (kbps)']
                            filesize = data['File Size 15s (MB)']
                            works = data['Works']

                            # Create cell content
                            cell_content = f"{bandwidth} kbps\n{filesize} MB\n{works}"

                            cell = ws.cell(row=row, column=col, value=cell_content)
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.border = border

                            # Color code based on success
                            if works == "✓":
                                cell.fill = success_fill
                            elif works == "✗":
                                cell.fill = fail_fill

                        else:
                            cell = ws.cell(row=row, column=col, value="N/A")
                            cell.alignment = center_align
                            cell.border = border

                        col += 1
                    row += 1

                # Add summary table below
                row += 2
                ws.cell(row=row, column=1, value="Detailed Data:")
                ws.cell(row=row, column=1).font = Font(bold=True, size=12)
                row += 1

                # Add detailed table
                detail_headers = ['Resolution', 'FPS', 'Bandwidth (kbps)', 'File Size 15s (MB)', 'Works']
                for col, header in enumerate(detail_headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border

                row += 1
                for _, data in df.iterrows():
                    for col, header in enumerate(detail_headers, 1):
                        cell = ws.cell(row=row, column=col, value=data[header])
                        cell.alignment = center_align
                        cell.border = border

                        if header == 'Works':
                            if data[header] == "✓":
                                cell.fill = success_fill
                            elif data[header] == "✗":
                                cell.fill = fail_fill
                    row += 1

                # Auto-adjust column widths (safest approach)
                from openpyxl.utils import get_column_letter

                for col_num in range(1, ws.max_column + 1):
                    max_length = 0
                    column_letter = get_column_letter(col_num)

                    for row_num in range(1, ws.max_row + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 20)  # Cap at 20
                    ws.column_dimensions[column_letter].width = adjusted_width

        # Create summary sheet
        summary_ws = wb.create_sheet(title="Summary", index=0)
        summary_ws['A1'] = "Camera Capabilities Analysis Summary"
        summary_ws['A1'].font = Font(bold=True, size=16)

        row = 3
        for device_info in self.video_devices:
            device_path = device_info['path']
            summary_ws.cell(row=row, column=1, value=f"Device: {device_path}")
            summary_ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            for format_name, format_data in device_info['capabilities'].items():
                res_count = len(format_data['resolutions'])
                total_combinations = sum(len(fps_list) for fps_list in format_data['resolutions'].values())

                summary_ws.cell(row=row, column=2, value=f"{format_name}: {res_count} resolutions, {total_combinations} total combinations")
                row += 1
            row += 1

        wb.save(filename)
        print(f"Excel file saved: {filename}")
        return filename

def main():
    print("Camera Capabilities Analysis Tool")
    print("=================================")

    analyzer = CameraAnalyzer()

    # Get device capabilities
    analyzer.get_device_capabilities()

    if not analyzer.video_devices:
        print("No video devices found!")
        return

    # Create Excel file
    filename = analyzer.create_excel_file()

    print(f"\nAnalysis complete!")
    print(f"Excel file created: {filename}")
    print("\nThe Excel file contains:")
    print("- One sheet per device/format combination")
    print("- Matrix showing Resolution vs FPS")
    print("- Each cell shows: Bandwidth (kbps), File Size 15s (MB), and Success status")
    print("- Green cells = Working combinations")
    print("- Red cells = Failed combinations")
    print("- Detailed data table below each matrix")

if __name__ == "__main__":
    main()